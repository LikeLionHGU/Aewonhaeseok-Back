#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""review_queue.csv에 남은 항목을 사람이 눈으로 판정할 수 있게 엑셀로 묶는다.

컬럼명만 봐서는 판정할 수 없다. 그 컬럼에 **실제로 어떤 값이 들어 있는지**를 봐야
'구분'이 분기인지 날짜인지 알 수 있다. 그래서 원본 데이터를 값까지 함께 싣는다.

    python3 ontology/make_review_workbook.py

출력: output/검증자료/검증_대기열.xlsx
      output/검증자료/{원본파일}.csv   (UTF-8, 엑셀에서 한글 안 깨지는 사본)
"""

from __future__ import annotations

import csv
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SAMPLES = ROOT / "data" / "samples"
QUEUE = ROOT / "output" / "review_queue.csv"
OUT = ROOT / "output" / "검증자료"

HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
TARGET_FILL = PatternFill("solid", fgColor="FFC000")   # 판정 대상 컬럼
ANSWER_FILL = PatternFill("solid", fgColor="FFF2CC")   # 사람이 채우는 칸
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ENCODINGS = ("utf-8-sig", "cp949", "euc-kr")


def read_rows(path: Path) -> list[list[str]]:
    for enc in ENCODINGS:
        try:
            with path.open(encoding=enc) as f:
                return [r for r in csv.reader(f)]
        except (UnicodeDecodeError, LookupError):
            continue
    raise SystemExit(f"인코딩 인식 실패: {path}")


def summarize(rows: list[list[str]], col: str) -> dict:
    """판정에 필요한 것은 컬럼명이 아니라 그 안의 값이다."""
    if not rows:
        return {}
    header = rows[0]
    if col not in header:
        return {"오류": f"'{col}' 컬럼 없음"}
    i = header.index(col)
    vals = [r[i].strip() for r in rows[1:] if len(r) > i and r[i].strip()]
    uniq: list[str] = []
    for v in vals:
        if v not in uniq:
            uniq.append(v)
    return {
        "행 수": len(vals),
        "고유값 수": len(uniq),
        # 키 이름이 "고유값 수"와 겹치면 아래 조회에서 개수를 집어 온다. 접두를 다르게 둔다.
        "값 목록": ", ".join(uniq[:12]) + (" …" if len(uniq) > 12 else ""),
        "앞 5개 값": ", ".join(vals[:5]),
        "값이 전부 다른가": "예 (행마다 고유 → 일련번호일 가능성)" if len(uniq) == len(vals)
                          else f"아니오 ({len(uniq)}종이 반복 → 분류값일 가능성)",
    }


def add_data_sheet(wb: Workbook, title: str, rows: list[list[str]], target: str) -> None:
    ws = wb.create_sheet(title[:31])
    for r in rows:
        ws.append(r)
    header = rows[0] if rows else []
    ti = header.index(target) if target in header else -1

    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = TARGET_FILL if (c - 1) == ti else HEAD_FILL
        cell.font = Font(bold=True, size=10, color="000000" if (c - 1) == ti else "FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = 22 if (c - 1) == ti else 15
    ws.row_dimensions[1].height = 32

    if ti >= 0:
        for r in range(2, len(rows) + 1):
            ws.cell(row=r, column=ti + 1).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.freeze_panes = "A2"


def main() -> int:
    if not QUEUE.exists():
        raise SystemExit(f"먼저 run.py를 돌려 대기열을 만드세요: {QUEUE}")

    with QUEUE.open(encoding="utf-8-sig") as f:
        queue = list(csv.DictReader(f))
    if not queue:
        print("대기열이 비어 있습니다. 판정할 것이 없습니다.")
        return 0

    OUT.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "판정"

    headers = ["번호", "출처파일", "원본 컬럼명", "후보 표준코드", "점수",
               "실제 값 요약", "고유값", "사람판정", "판정자", "판정 근거"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    widths = [6, 34, 16, 14, 7, 46, 40, 14, 12, 46]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    seen_files: set[str] = set()
    for n, q in enumerate(queue, start=1):
        src = q["출처파일"]
        col = q["원본명"]
        path = SAMPLES / src
        info = summarize(read_rows(path), col) if path.exists() else {"오류": "원본 없음"}

        ws.append([
            n, src, col, q.get("후보 표준코드", ""), q.get("점수", ""),
            f"{info.get('행 수', '')}행 / 고유값 {info.get('고유값 수', '')}개 · "
            f"{info.get('값이 전부 다른가', '')}",
            info.get("값 목록", ""),
            "", "", "",
        ])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=n + 1, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)
            if headers[c - 1] in ("사람판정", "판정자", "판정 근거"):
                cell.fill = ANSWER_FILL
        ws.row_dimensions[n + 1].height = 46

        if path.exists() and src not in seen_files:
            seen_files.add(src)
            add_data_sheet(wb, Path(src).stem, read_rows(path), col)
            # 엑셀에서 한글이 깨지지 않는 UTF-8 사본도 함께 둔다 (원본은 CP949)
            rows = read_rows(path)
            with (OUT / src).open("w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerows(rows)

    ws.freeze_panes = "A2"

    guide = wb.create_sheet("판정 방법")
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 108
    for a, b in [
        ("무엇을 정하는가", "이 컬럼이 표준 사전의 어느 개념을 가리키는지 정한다. "
                        "컬럼명만 보지 말고 '실제 값 요약'과 데이터 시트의 노란 열을 볼 것."),
        ("", ""),
        ("판정 칸에 적을 것", ""),
        ("  표준코드", "예: MD-006. 그 코드의 동의어로 등록된다."),
        ("  승인", "'후보 표준코드'가 맞으면 이렇게 적는다."),
        ("  기각", "어느 표준에도 해당하지 않으면. 사전에 넣지 않는다."),
        ("  빈칸", "기각과 같게 처리된다(반영 안 됨)."),
        ("", ""),
        ("판정 기준", ""),
        ("  값이 행마다 고유", "일련번호일 가능성이 높다. 파일 판본이 바뀌어도 같은 대상이 "
                          "같은 번호를 유지하는지 확인할 것. 유지되면 식별자, 아니면 단순 행번호."),
        ("  값이 몇 종 반복", "분류값이다. 그 종류를 보고 개념을 정한다(분기/등급/구분 등)."),
        ("  값이 날짜 형식", "시간 축 메타다. 측정일자인지 자료기준일자인지 구분할 것."),
        ("", ""),
        ("주의 — 같은 이름 다른 뜻", "같은 컬럼명이 파일마다 다른 것을 가리키면 전역 사전에 넣으면 안 된다. "
                            "한쪽을 맞추면 다른 쪽이 반드시 틀린다. 이런 건 기각하고 "
                            "파일별 규칙으로 따로 처리해야 한다."),
        ("", ""),
        ("판정 후", "이 시트의 판정을 output/review_queue.csv의 '사람판정'·'판정자' 칸에 옮긴 뒤:"),
        ("", "  python3 pipeline/apply_review.py output/review_queue.csv"),
        ("", "  python3 ontology/import_review_log.py"),
        ("", "  python3 ontology/export_ontology.py"),
    ]:
        guide.append([a, b])
    for row in guide.iter_rows():
        row[0].font = Font(bold=True, size=11)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[1].font = Font(size=10)

    path = OUT / "검증_대기열.xlsx"
    wb.save(path)
    print(f"저장: {path}")
    print(f"      원본 UTF-8 사본 {len(seen_files)}개도 같은 폴더에 두었습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
